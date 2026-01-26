#!/usr/bin/env python3
"""
Generate English chapter files by:
1. Reading Spanish capitulo_XX.html files as templates
2. Extracting English translations from Excel
3. Creating chapter_XX.html files with matching structure
"""

import os
import re
from openpyxl import load_workbook


def extract_english_translations(excel_path):
    """Extract English translations organized by chapter."""
    wb = load_workbook(excel_path)
    translations = {}

    for sheet_name in wb.sheetnames:
        cap_match = re.match(r'^cap (\d+)$', sheet_name)
        if not cap_match:
            continue

        chapter_num = int(cap_match.group(1))
        sheet = wb[sheet_name]

        # Extract all non-empty text from column A
        content = []
        for row in sheet.iter_rows(values_only=True):
            cell = row[0] if row else None
            if cell is not None and str(cell).strip():
                text = str(cell).strip()
                # Skip intro text and Spanish content
                if any(skip in text.lower() for skip in ['to achieve better', 'each chapter has', 'some questions have', 'the questions will', 'wishing you']):
                    continue
                # Skip Spanish text
                if '¿' in text or 'según' in text.lower() or 'escoja' in text.lower():
                    continue
                content.append(text)

        translations[chapter_num] = content

    return translations


def get_chapter_title(chapter_num, translations):
    """Get the English chapter title from translations."""
    content = translations.get(chapter_num, [])
    for line in content:
        # Look for chapter title pattern like "1. Understanding Our Differences"
        match = re.match(rf'^{chapter_num}\.\s+(.+)$', line)
        if match:
            return match.group(1)
    return f"Chapter {chapter_num}"


def translate_capitulo(spanish_html, chapter_num, translations):
    """Translate a Spanish capitulo HTML to English chapter HTML."""
    html = spanish_html
    content = translations.get(chapter_num, [])

    # Build lookup of English questions and options
    english_questions = []
    english_options = {}
    current_q = None

    for line in content:
        # Check if it's a question (ends with ?)
        if line.endswith('?') or line.startswith('Which') or line.startswith('What') or line.startswith('How') or line.startswith('Why') or line.startswith('Is ') or line.startswith('Are ') or line.startswith('Do ') or line.startswith('Did ') or line.startswith('Can ') or line.startswith('Would ') or line.startswith('When ') or line.startswith('For ') or line.startswith('On a scale') or line.startswith('Check all') or line.startswith('Select the') or 'usually' in line:
            current_q = line
            english_questions.append(line)
            english_options[line] = []
        # Check if it's an option
        elif current_q:
            # Bullet point (• Yes, • No)
            bullet_match = re.match(r'^[•·]\s*(.+)$', line)
            if bullet_match:
                english_options[current_q].append(bullet_match.group(1).strip())
                continue
            # Letter option (A. Money, B. Happiness)
            letter_match = re.match(r'^([A-D])\.\s*(.+)$', line)
            if letter_match:
                english_options[current_q].append(letter_match.group(2).strip())
                continue
            # Checkbox (☐ Option)
            checkbox_match = re.match(r'^[☐□]\s*(.+)$', line)
            if checkbox_match:
                english_options[current_q].append(checkbox_match.group(1).strip())
                continue
            # True/False
            if line in ['True', 'False', 'Yes', 'No']:
                english_options[current_q].append(line)
                continue

    # Get discussion prompts (lines after "Discuss:" that aren't questions/options)
    discussion_prompts = []
    in_discussion = False
    for line in content:
        if line.lower().startswith('discuss') or line.lower().startswith('reflection') or line.lower().startswith('share'):
            in_discussion = True
            continue
        if in_discussion and not line.endswith('?') and not line.startswith('•') and not re.match(r'^[A-D]\.', line):
            if not any(skip in line.lower() for skip in ['chapter', 'husband', 'wife']) or len(line) > 20:
                discussion_prompts.append(line)

    # Change language attribute
    html = html.replace('lang="es"', 'lang="en"')

    # Update title
    html = re.sub(r'<title>Capítulo (\d+)[^<]*</title>',
                  f'<title>Chapter {chapter_num} - After watching the video</title>', html)

    # Update h1
    html = re.sub(r'<h1>[^<]*</h1>',
                  '<h1>After watching the video, complete this exercise:</h1>', html)

    # Update chapter name in JavaScript
    html = re.sub(r'window\.chapterName = "[^"]*"',
                  f'window.chapterName = "Chapter {chapter_num}"', html)

    # Replace question titles
    for i, eng_q in enumerate(english_questions[:20], 1):  # Limit to first 20 questions
        # Find Q{i}: in the HTML and replace the question text
        # First try with exact Q{i} pattern
        pattern = rf'(<div class="question-title">Q{i}:\s*)[^<]+'
        # Escape special regex chars in the English question
        safe_q = eng_q.replace('\\', '\\\\').replace('"', '\\"')
        replacement = f'\\1{eng_q}'
        html = re.sub(pattern, replacement, html, count=1)

    # Replace option labels (Spanish -> English)
    option_replacements = {
        'Si': 'Yes',
        'No': 'No',
        'Verdadero': 'True',
        'Falso': 'False',
        'Dinero': 'Money',
        'Felicidad': 'Happiness',
        'Paz': 'Peace',
        'Amor': 'Love',
        'Autoestima': 'Self-esteem',
        'La infancia': 'Childhood',
        'La adolescencia': 'Adolescence',
        'La adultez': 'Adulthood',
        'Todas las anteriores': 'All of the above',
    }

    for spanish, english in option_replacements.items():
        # Replace in labels
        html = re.sub(rf'>({spanish})<', f'>{english}<', html)
        # Replace in values
        html = re.sub(rf'value="{spanish}"', f'value="{english}"', html)

    # Replace UI text
    ui_replacements = {
        'Conversen:': 'Discuss:',
        'Correo Electrónico': 'Email Address',
        'Escribe tu respuesta aquí...': 'Write your answer here...',
        'ejemplo@correo.com': 'example@email.com',
        'Enviar Encuesta': 'Submit Survey',
        'Limpiar': 'Clear',
        '¡Encuesta enviada correctamente!': 'Survey submitted successfully!',
        'Resumen de tus respuestas': 'Summary of your answers',
        'Reflexión:': 'Reflection:',
    }

    for spanish, english in ui_replacements.items():
        html = html.replace(spanish, english)

    # Update surveyQuestions in JavaScript
    for i, eng_q in enumerate(english_questions[:20], 1):
        # Escape quotes for JavaScript
        safe_q = eng_q.replace('"', '\\"')
        pattern = rf'(q{i}:\s*")[^"]*(")'
        replacement = f'\\1Q{i}: {safe_q}\\2'
        html = re.sub(pattern, replacement, html, count=1)

    # Try to replace conversation/discussion prompts
    if discussion_prompts:
        for prompt in discussion_prompts[:2]:
            # Escape for HTML/JS
            safe_prompt = prompt.replace('"', '\\"')
            # Replace Spanish conversation prompts with English
            # This is harder to do generically, but try common patterns
            pass

    return html


def main():
    excel_path = 'perfecto  imperfecto matrimonio done questions english.xlsx'
    translations = extract_english_translations(excel_path)

    for chapter_num in range(1, 23):
        capitulo_path = f'capitulo_{chapter_num:02d}.html'

        if not os.path.exists(capitulo_path):
            print(f"Skipping chapter {chapter_num}: {capitulo_path} not found")
            continue

        print(f"Processing Chapter {chapter_num}...")

        with open(capitulo_path, 'r', encoding='utf-8') as f:
            spanish_html = f.read()

        english_html = translate_capitulo(spanish_html, chapter_num, translations)

        output_path = f'chapter_{chapter_num:02d}.html'
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(english_html)

        print(f"  Created {output_path}")

    print("\nDone!")


if __name__ == '__main__':
    main()
