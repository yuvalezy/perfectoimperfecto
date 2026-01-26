#!/usr/bin/env python3
"""
Generate English chapter files from Excel translations and Spanish capitulo templates.
"""

import os
import re
from openpyxl import load_workbook

# English UI translations
UI_TRANSLATIONS = {
    'Después de ver el video, realiza este ejercicio:': 'After watching the video, complete this exercise:',
    'Conversen:': 'Discuss:',
    'Correo Electrónico': 'Email Address',
    'Escribe tu respuesta aquí...': 'Write your answer here...',
    'ejemplo@correo.com': 'example@email.com',
    'Enviar Encuesta': 'Submit Survey',
    'Limpiar': 'Clear',
    '¡Encuesta enviada correctamente!': 'Survey submitted successfully!',
    'Resumen de tus respuestas': 'Summary of your answers',
    'Si': 'Yes',
    'No': 'No',
    'Capítulo': 'Chapter',
}

def parse_excel_chapter(sheet):
    """Parse a chapter sheet and extract questions, options, and discussion prompts."""
    content = []
    current_item = None

    for row in sheet.iter_rows(values_only=True):
        cell = row[0] if row else None
        if cell is None or (isinstance(cell, str) and cell.strip() == ''):
            continue

        cell = str(cell).strip()

        # Skip intro text
        if cell.startswith('To achieve better results') or cell.startswith('Each chapter has'):
            continue
        if cell.startswith('Some questions have') or cell.startswith('The questions will'):
            continue
        if cell.startswith('Wishing you much'):
            continue

        # Skip Spanish text in mixed sheets
        if any(spanish in cell.lower() for spanish in ['conversen:', '¿', 'según', 'escoja', 'hay que']):
            continue

        # Check if it's a chapter title (e.g., "1. Understanding Our Differences")
        chapter_match = re.match(r'^(\d+)\.\s+(.+)$', cell)
        if chapter_match:
            content.append({'type': 'chapter_title', 'text': cell})
            continue

        # Check if it's a numbered question or just a question
        if cell.endswith('?') or cell.startswith('Which') or cell.startswith('What') or cell.startswith('How') or cell.startswith('Why') or cell.startswith('Is ') or cell.startswith('Are ') or cell.startswith('Do ') or cell.startswith('Did ') or cell.startswith('Can ') or cell.startswith('Would ') or cell.startswith('When ') or cell.startswith('For '):
            if current_item:
                content.append(current_item)
            current_item = {'type': 'question', 'text': cell, 'options': []}
            continue

        # Check for various option formats
        # Bullet point options (• Option)
        bullet_match = re.match(r'^[•·]\s*(.+)$', cell)
        if bullet_match:
            option_text = bullet_match.group(1).strip()
            if current_item and current_item['type'] == 'question':
                current_item['options'].append(option_text)
            continue

        # Letter options (A. Option, B. Option)
        letter_match = re.match(r'^([A-D])\.\s*(.+)$', cell)
        if letter_match:
            option_text = letter_match.group(2).strip()
            if current_item and current_item['type'] == 'question':
                current_item['options'].append(option_text)
            continue

        # Checkbox options (☐ Option)
        checkbox_match = re.match(r'^[☐□]\s*(.+)$', cell)
        if checkbox_match:
            option_text = checkbox_match.group(1).strip()
            if current_item and current_item['type'] == 'question':
                current_item['options'].append(option_text)
                current_item['checkbox'] = True
            continue

        # Scale (1 2 3 4 5 6 7 8 9 10)
        if re.match(r'^1\s+2\s+3\s+4\s+5', cell):
            if current_item:
                current_item['scale'] = True
            continue

        # Discussion/Reflection prompts
        if cell.lower().startswith('discuss') or cell.lower().startswith('reflection') or cell.lower().startswith('share'):
            if current_item:
                content.append(current_item)
            current_item = {'type': 'discussion', 'text': cell, 'prompts': []}
            continue

        # Check for section headers like "For Men", "For Women", "Examples:"
        if cell in ['For Men', 'For Women', 'Examples:', 'Select the correct answers:', 'Check all that apply:']:
            if current_item:
                content.append(current_item)
            current_item = {'type': 'section', 'text': cell}
            continue

        # Otherwise, it might be a discussion prompt or continuation
        if current_item:
            if current_item['type'] == 'discussion':
                current_item['prompts'].append(cell)
            elif current_item['type'] == 'question' and not current_item.get('options'):
                # It's a question without standard options - might be discussion text
                current_item['prompts'] = current_item.get('prompts', [])
                current_item['prompts'].append(cell)

    if current_item:
        content.append(current_item)

    return content


def read_capitulo_file(chapter_num):
    """Read the Spanish capitulo file and extract its structure."""
    filename = f'capitulo_{chapter_num:02d}.html'
    filepath = os.path.join(os.path.dirname(os.path.dirname(__file__)), filename)

    if not os.path.exists(filepath):
        return None

    with open(filepath, 'r', encoding='utf-8') as f:
        return f.read()


def generate_english_chapter(chapter_num, excel_content, spanish_html):
    """Generate English chapter HTML based on Spanish template and Excel content."""

    # Start with Spanish HTML and modify it
    html = spanish_html

    # Change language
    html = html.replace('lang="es"', 'lang="en"')

    # Update title
    html = re.sub(r'<title>Capítulo (\d+)', f'<title>Chapter {chapter_num}', html)

    # Update h1
    html = html.replace(
        'Después de ver el video, realiza este ejercicio:',
        'After watching the video, complete this exercise:'
    )

    # Update chapter name in JavaScript
    html = re.sub(
        r'window\.chapterName = "Capítulo \d+"',
        f'window.chapterName = "Chapter {chapter_num}"',
        html
    )

    # Update UI elements
    html = html.replace('Conversen:', 'Discuss:')
    html = html.replace('>Correo Electrónico<', '>Email Address<')
    html = html.replace('placeholder="Escribe tu respuesta aquí..."', 'placeholder="Write your answer here..."')
    html = html.replace('placeholder="ejemplo@correo.com"', 'placeholder="example@email.com"')
    html = html.replace('>Enviar Encuesta<', '>Submit Survey<')
    html = html.replace('>Limpiar<', '>Clear<')
    html = html.replace('¡Encuesta enviada correctamente!', 'Survey submitted successfully!')
    html = html.replace('Resumen de tus respuestas', 'Summary of your answers')

    # Extract questions and options from Excel content
    questions = [item for item in excel_content if item['type'] == 'question']
    discussions = [item for item in excel_content if item['type'] == 'discussion']

    # Replace question titles
    q_num = 1
    for question in questions:
        if question.get('options'):
            # Find the Q{n} pattern and replace the text
            pattern = rf'(Q{q_num}:\s*)[^<]+'
            replacement = f'Q{q_num}: {question["text"]}'
            html = re.sub(pattern, replacement, html, count=1)

            # Also update surveyQuestions
            pattern = rf'(q{q_num}:\s*")[^"]+'
            replacement = f'q{q_num}: "Q{q_num}: {question["text"]}'
            html = re.sub(pattern, replacement, html, count=1)

            q_num += 1

    # Replace option labels
    # This is tricky because options are in specific positions
    # We need to match Spanish labels and replace with English

    # Common option replacements
    html = re.sub(r'>Si<', '>Yes<', html)
    html = re.sub(r'"Si"', '"Yes"', html)
    html = re.sub(r'"No"', '"No"', html)

    # Replace discussion prompts
    for discussion in discussions:
        if discussion.get('prompts'):
            for prompt in discussion['prompts']:
                # Try to find corresponding Spanish text and replace
                pass

    return html


def create_minimal_english_chapter(chapter_num, excel_content):
    """Create a minimal English chapter when the structure differs too much."""

    questions = [item for item in excel_content if item['type'] == 'question']
    discussions = [item for item in excel_content if item['type'] == 'discussion']

    # Build question sections HTML
    question_html = []
    survey_questions = []
    q_idx = 1

    for q in questions:
        if q.get('options'):
            options_html = []
            for i, opt in enumerate(q['options'], 1):
                opt_id = f'q{q_idx}_{i}'
                required = ' required' if i == 1 else ''
                options_html.append(f'''                    <div class="option">
                        <input type="radio" id="{opt_id}" name="q{q_idx}" value="{opt}"{required}>
                        <label for="{opt_id}">{opt}</label>
                    </div>''')

            question_html.append(f'''            <!-- Q{q_idx} -->
            <div class="question-section">
                <div class="question-title">Q{q_idx}: {q['text']}</div>
                <div class="options">
{chr(10).join(options_html)}
                </div>
            </div>
''')
            survey_questions.append(f'            q{q_idx}: "Q{q_idx}: {q["text"]}"')
            q_idx += 1

    # Discussion section
    discussion_html = ''
    if discussions:
        prompts = []
        for d in discussions:
            for p in d.get('prompts', []):
                prompts.append(p)

        if prompts:
            discussion_html = f'''            <!-- Conversation Section -->
            <div class="question-section">
                <div class="question-title">Discuss:</div>
                <div class="text-input-section">
                    <label for="conversation" style="color: #666; font-size: 14px;">{prompts[0] if prompts else 'Share your thoughts.'}</label>
                    <textarea id="conversation" name="conversation" placeholder="Write your answer here..." required></textarea>
                    <div class="error-message" id="textError" style="display: none;"></div>
                </div>
            </div>
'''
            survey_questions.append(f'            conversation: "{prompts[0] if prompts else "Share your thoughts."}"')

    html = f'''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Chapter {chapter_num} - After watching the video</title>
    <link rel="stylesheet" href="survey.css">
</head>
<body>
    <div class="container">
        <h1>After watching the video, complete this exercise:</h1>
        <form id="pollForm">
{chr(10).join(question_html)}
{discussion_html}
            <!-- Email Section -->
            <div class="question-section">
                <div class="question-title"><label for="email">Email Address</label></div>
                <div class="text-input-section">
                    <input type="email" id="email" name="email" placeholder="example@email.com" required>
                    <div class="error-message" id="emailError" style="display: none;"></div>
                </div>
            </div>

            <!-- reCAPTCHA -->
            <div class="question-section">
                <div class="g-recaptcha" data-sitekey="6LeyoggsAAAAAAgXzEg9PAC9ypZtr-yyc24cAnN_"></div>
                <div class="error-message" id="recaptchaError" style="display: none;"></div>
            </div>

            <!-- Buttons -->
            <div class="button-group">
                <button type="submit" class="submit-btn" id="submitBtn">Submit Survey</button>
                <button type="reset" class="reset-btn" id="resetBtn">Clear</button>
            </div>
        </form>

        <!-- Success Message -->
        <div class="success-message" id="successMessage">
            Survey submitted successfully!
        </div>

        <!-- Summary Section -->
        <div class="summary-section" id="summarySection" style="display: none;">
            <h2 class="summary-title">Summary of your answers</h2>
            <div id="summaryContent"></div>
        </div>
    </div>

    <script>
        // Define questions for this survey
        window.surveyQuestions = {{
{(","+chr(10)).join(survey_questions)}
        }};

        // Define the chapter name for this survey
        window.chapterName = "Chapter {chapter_num}";
    </script>

    <!-- EmailJS Library -->
    <script src="https://cdn.jsdelivr.net/npm/emailjs-com@3/dist/email.min.js"></script>
    <script>
        (function() {{
            emailjs.init("wZ_Z4F9Y-8CcFzD2g");
        }})();
    </script>

    <!-- Google reCAPTCHA v2 -->
    <script src="https://www.google.com/recaptcha/api.js" async defer></script>

    <script src="survey.js?v=20250110005"></script>
</body>
</html>
'''
    return html


def main():
    # Load the Excel file
    excel_path = 'perfecto  imperfecto matrimonio done questions english.xlsx'
    wb = load_workbook(excel_path)

    # Process each chapter sheet
    for sheet_name in wb.sheetnames:
        # Only process chapter sheets
        cap_match = re.match(r'^cap (\d+)$', sheet_name)
        if not cap_match:
            continue

        chapter_num = int(cap_match.group(1))
        print(f"Processing Chapter {chapter_num}...")

        sheet = wb[sheet_name]
        excel_content = parse_excel_chapter(sheet)

        if not excel_content:
            print(f"  Warning: No content found for Chapter {chapter_num}")
            continue

        # Print what we found
        questions = [item for item in excel_content if item['type'] == 'question']
        print(f"  Found {len(questions)} questions")

        # Generate English HTML
        html = create_minimal_english_chapter(chapter_num, excel_content)

        # Write the file
        output_path = f'chapter_{chapter_num:02d}.html'
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(html)
        print(f"  Created {output_path}")

    print("\nDone! Now update index.html to include links to English chapters.")


if __name__ == '__main__':
    main()
