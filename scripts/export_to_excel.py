#!/usr/bin/env python3
"""
Export survey chapters to Excel for easy translation/editing.
Extracts all translatable content from capitulo_*.html files.
"""

import os
import re
import glob
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Configuration
HTML_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUTPUT_FILE = os.path.join(HTML_DIR, 'survey_content.xlsx')


def extract_chapter_data(html_path):
    """Extract all translatable content from a chapter HTML file."""
    with open(html_path, 'r', encoding='utf-8') as f:
        soup = BeautifulSoup(f.read(), 'html.parser')

    # Get chapter number from filename
    filename = os.path.basename(html_path)
    chapter_match = re.search(r'capitulo_(\d+)', filename)
    chapter_num = int(chapter_match.group(1)) if chapter_match else 0

    data = {
        'chapter_num': chapter_num,
        'filename': filename,
        'title': '',
        'main_instruction': '',
        'questions': [],
        'conversation_label': '',
        'conversation_placeholder': '',
        'email_label': '',
        'email_placeholder': '',
        'submit_button': '',
        'reset_button': '',
        'success_message': '',
        'summary_title': ''
    }

    # Extract title
    title_tag = soup.find('title')
    if title_tag:
        data['title'] = title_tag.get_text(strip=True)

    # Extract main instruction (h1)
    h1_tag = soup.find('h1')
    if h1_tag:
        data['main_instruction'] = h1_tag.get_text(strip=True)

    # Extract questions
    question_sections = soup.find_all('div', class_='question-section')
    for section in question_sections:
        question_title = section.find('div', class_='question-title')
        if not question_title:
            continue

        title_text = question_title.get_text(strip=True)

        # Check if this is a regular question (Q1, Q2, etc.)
        if title_text.startswith('Q'):
            options_div = section.find('div', class_='options')
            options = []
            if options_div:
                for option in options_div.find_all('div', class_='option'):
                    label = option.find('label')
                    if label:
                        options.append(label.get_text(strip=True))

            # Get question ID from input name
            input_tag = section.find('input')
            q_id = input_tag.get('name', '') if input_tag else ''

            data['questions'].append({
                'id': q_id,
                'text': title_text,
                'options': options
            })

        # Check for conversation section
        elif 'Convers' in title_text:
            text_section = section.find('div', class_='text-input-section')
            if text_section:
                label = text_section.find('label')
                if label:
                    data['conversation_label'] = label.get_text(strip=True)
                textarea = text_section.find('textarea')
                if textarea and textarea.get('placeholder'):
                    data['conversation_placeholder'] = textarea.get('placeholder')

        # Check for email section
        elif 'Correo' in title_text or 'Email' in title_text:
            text_section = section.find('div', class_='text-input-section')
            if text_section:
                input_tag = text_section.find('input', {'type': 'email'})
                if input_tag and input_tag.get('placeholder'):
                    data['email_placeholder'] = input_tag.get('placeholder')
            data['email_label'] = title_text

    # Extract button texts
    submit_btn = soup.find('button', class_='submit-btn')
    if submit_btn:
        data['submit_button'] = submit_btn.get_text(strip=True)

    reset_btn = soup.find('button', class_='reset-btn')
    if reset_btn:
        data['reset_button'] = reset_btn.get_text(strip=True)

    # Extract success message
    success_msg = soup.find('div', class_='success-message')
    if success_msg:
        data['success_message'] = success_msg.get_text(strip=True)

    # Extract summary title
    summary_title = soup.find('h2', class_='summary-title')
    if summary_title:
        data['summary_title'] = summary_title.get_text(strip=True)

    return data


def create_excel(chapters_data):
    """Create Excel workbook with all chapter content."""
    wb = Workbook()

    # Styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    chapter_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    wrap_alignment = Alignment(wrap_text=True, vertical='top')

    # === Sheet 1: Questions ===
    ws_questions = wb.active
    ws_questions.title = 'Questions'

    # Find max options across all chapters
    max_options = 0
    for ch in chapters_data:
        for q in ch['questions']:
            max_options = max(max_options, len(q['options']))

    # Headers
    headers = ['Chapter', 'Question ID', 'Question Text'] + [f'Option {i+1}' for i in range(max_options)]
    for col, header in enumerate(headers, 1):
        cell = ws_questions.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Data
    row = 2
    for ch in sorted(chapters_data, key=lambda x: x['chapter_num']):
        chapter_start_row = row
        for q in ch['questions']:
            ws_questions.cell(row=row, column=1, value=ch['chapter_num']).border = thin_border
            ws_questions.cell(row=row, column=2, value=q['id']).border = thin_border
            cell = ws_questions.cell(row=row, column=3, value=q['text'])
            cell.border = thin_border
            cell.alignment = wrap_alignment

            for opt_idx, opt in enumerate(q['options']):
                cell = ws_questions.cell(row=row, column=4 + opt_idx, value=opt)
                cell.border = thin_border
                cell.alignment = wrap_alignment

            row += 1

        # Add light fill for chapter grouping
        for r in range(chapter_start_row, row):
            if ch['chapter_num'] % 2 == 0:
                ws_questions.cell(row=r, column=1).fill = chapter_fill

    # Column widths
    ws_questions.column_dimensions['A'].width = 10
    ws_questions.column_dimensions['B'].width = 12
    ws_questions.column_dimensions['C'].width = 60
    for i in range(max_options):
        ws_questions.column_dimensions[get_column_letter(4 + i)].width = 40

    # === Sheet 2: UI Elements ===
    ws_ui = wb.create_sheet('UI Elements')

    ui_headers = ['Chapter', 'Title', 'Main Instruction', 'Conversation Label',
                  'Conversation Placeholder', 'Email Label', 'Email Placeholder',
                  'Submit Button', 'Reset Button', 'Success Message', 'Summary Title']

    for col, header in enumerate(ui_headers, 1):
        cell = ws_ui.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    row = 2
    for ch in sorted(chapters_data, key=lambda x: x['chapter_num']):
        values = [
            ch['chapter_num'],
            ch['title'],
            ch['main_instruction'],
            ch['conversation_label'],
            ch['conversation_placeholder'],
            ch['email_label'],
            ch['email_placeholder'],
            ch['submit_button'],
            ch['reset_button'],
            ch['success_message'],
            ch['summary_title']
        ]
        for col, val in enumerate(values, 1):
            cell = ws_ui.cell(row=row, column=col, value=val)
            cell.border = thin_border
            cell.alignment = wrap_alignment
        row += 1

    # Column widths for UI sheet
    widths = [10, 40, 50, 60, 30, 25, 25, 20, 15, 35, 30]
    for i, w in enumerate(widths):
        ws_ui.column_dimensions[get_column_letter(i + 1)].width = w

    # Freeze first row in both sheets
    ws_questions.freeze_panes = 'A2'
    ws_ui.freeze_panes = 'A2'

    wb.save(OUTPUT_FILE)
    return OUTPUT_FILE


def main():
    # Find all chapter files
    pattern = os.path.join(HTML_DIR, 'capitulo_*.html')
    html_files = sorted(glob.glob(pattern))

    if not html_files:
        print(f"No chapter files found in {HTML_DIR}")
        return

    print(f"Found {len(html_files)} chapter files")

    # Extract data from all chapters
    chapters_data = []
    for html_file in html_files:
        print(f"Processing: {os.path.basename(html_file)}")
        data = extract_chapter_data(html_file)
        chapters_data.append(data)

    # Create Excel file
    output_path = create_excel(chapters_data)
    print(f"\nExported to: {output_path}")
    print(f"Total chapters: {len(chapters_data)}")
    print(f"Total questions: {sum(len(ch['questions']) for ch in chapters_data)}")


if __name__ == '__main__':
    main()
