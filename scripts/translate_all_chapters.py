#!/usr/bin/env python3
"""
Translate all chapter files using Excel content.
Processes each sheet, extracts questions/options, and updates chapter files.
"""

import os
import re
from openpyxl import load_workbook


def extract_chapter_content(sheet):
    """Extract structured content from a chapter sheet."""
    content = {
        'questions': [],
        'discussion_prompts': []
    }

    current_question = None
    in_discussion = False

    for row in sheet.iter_rows(values_only=True):
        cell = row[0] if row else None
        if cell is None:
            continue

        text = str(cell).strip()
        if not text:
            continue

        # Skip intro text
        if text.startswith('To achieve better') or text.startswith('Each chapter'):
            continue

        # Check for chapter title (e.g., "4. Understanding Women's Needs")
        if re.match(r'^\d+\.\s+', text):
            continue

        # Check for discussion section
        if text.lower().startswith('discuss'):
            in_discussion = True
            continue

        if in_discussion:
            # This is a discussion prompt
            if text and not text.startswith('•'):
                content['discussion_prompts'].append(text)
            continue

        # Check if it's a question (ends with ?)
        if text.endswith('?') or text.endswith(':'):
            if current_question and current_question['options']:
                content['questions'].append(current_question)
            current_question = {'text': text, 'options': []}
            continue

        # Check if it's an option
        if text.startswith('•') or text.startswith('·'):
            option = text[1:].strip()
            if current_question:
                current_question['options'].append(option)
        elif re.match(r'^[A-D]\.\s', text):
            option = re.sub(r'^[A-D]\.\s*', '', text).strip()
            if current_question:
                current_question['options'].append(option)
        elif text.startswith('☐') or text.startswith('□'):
            option = text[1:].strip()
            if current_question:
                current_question['options'].append(option)

    # Don't forget the last question
    if current_question and current_question['options']:
        content['questions'].append(current_question)

    return content


def update_chapter_file(chapter_num, content):
    """Update a chapter file with English content."""
    filepath = f'chapter_{chapter_num:02d}.html'

    if not os.path.exists(filepath):
        print(f"  File not found: {filepath}")
        return False

    with open(filepath, 'r', encoding='utf-8') as f:
        html = f.read()

    original_html = html
    questions = content['questions']
    discussion_prompts = content['discussion_prompts']

    # Update each question's options
    for q_idx, question in enumerate(questions, 1):
        # Find the question section in HTML
        # Update option values and labels
        for opt_idx, option in enumerate(question['options'], 1):
            # Pattern to match the option input and label
            opt_id = f'q{q_idx}_{opt_idx}'

            # Find and replace the value attribute
            pattern = rf'(<input[^>]*id="{opt_id}"[^>]*value=")[^"]*(")'
            html = re.sub(pattern, rf'\1{option}\2', html)

            # Find and replace the label text
            pattern = rf'(<label for="{opt_id}">)[^<]*(</label>)'
            html = re.sub(pattern, rf'\1{option}\2', html)

    # Update discussion prompts
    if discussion_prompts:
        for i, prompt in enumerate(discussion_prompts[:2]):
            if i == 0:
                # First conversation field
                pattern = r'(<label for="conversation"[^>]*>)[^<]*(</label>)'
                html = re.sub(pattern, rf'\1{prompt}\2', html)
                pattern = r'(conversation:\s*")[^"]*(")'
                html = re.sub(pattern, rf'\1{prompt}\2', html)
            else:
                # Second conversation field
                pattern = r'(<label for="conversation2"[^>]*>)[^<]*(</label>)'
                html = re.sub(pattern, rf'\1{prompt}\2', html)
                pattern = r'(conversation2:\s*")[^"]*(")'
                html = re.sub(pattern, rf'\1{prompt}\2', html)

    if html != original_html:
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(html)
        return True
    return False


def main():
    excel_path = 'perfecto  imperfecto matrimonio done questions english.xlsx'
    wb = load_workbook(excel_path)

    for chapter_num in range(1, 23):
        sheet_name = f'cap {chapter_num}'
        if sheet_name not in wb.sheetnames:
            print(f"Sheet not found: {sheet_name}")
            continue

        print(f"Processing Chapter {chapter_num}...")
        sheet = wb[sheet_name]
        content = extract_chapter_content(sheet)

        print(f"  Found {len(content['questions'])} questions, {len(content['discussion_prompts'])} discussion prompts")

        if update_chapter_file(chapter_num, content):
            print(f"  Updated chapter_{chapter_num:02d}.html")
        else:
            print(f"  No changes needed for chapter_{chapter_num:02d}.html")


if __name__ == '__main__':
    main()
