#!/usr/bin/env python3
"""
Import survey content from Excel and generate HTML chapter files.
Supports multiple languages by generating different file sets.

Usage:
    python import_from_excel.py                    # Generate Spanish (default)
    python import_from_excel.py --lang en          # Generate English
    python import_from_excel.py --lang en --prefix chapter  # English with custom prefix
"""

import os
import argparse
from openpyxl import load_workbook

# Configuration
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
HTML_DIR = os.path.dirname(SCRIPT_DIR)
DEFAULT_INPUT_FILE = os.path.join(HTML_DIR, 'survey_content.xlsx')

# Language-specific defaults
LANG_CONFIG = {
    'es': {
        'prefix': 'capitulo',
        'html_lang': 'es',
        'css_file': 'survey.css',
        'js_file': 'survey.js?v=20250110005'
    },
    'en': {
        'prefix': 'chapter',
        'html_lang': 'en',
        'css_file': 'survey.css',
        'js_file': 'survey.js?v=20250110005'
    }
}

HTML_TEMPLATE = '''<!DOCTYPE html>
<html lang="{html_lang}">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{title}</title>
    <link rel="stylesheet" href="{css_file}">
</head>
<body>
    <div class="container">
        <h1>{main_instruction}</h1>
        <form id="pollForm">
{questions_html}
            <!-- Conversation Section -->
            <div class="question-section">
                <div class="question-title">{conversation_title}</div>
                <div class="text-input-section">
                    <label for="conversation" style="color: #666; font-size: 14px;">{conversation_label}</label>
                    <textarea id="conversation" name="conversation" placeholder="{conversation_placeholder}" required></textarea>
                    <div class="error-message" id="textError" style="display: none;"></div>
                </div>
            </div>

            <!-- Email Section -->
            <div class="question-section">
                <div class="question-title"><label for="email">{email_label}</label></div>
                <div class="text-input-section">
                    <input type="email" id="email" name="email" placeholder="{email_placeholder}" required>
                    <div class="error-message" id="emailError" style="display: none;"></div>
                </div>
            </div>

            <!-- reCAPTCHA -->
            <div class="question-section">
                <div class="g-recaptcha" data-sitekey="6LeyoggsAAAAAAgXzEg9PAC9ypZtr-yyc24cAnN_"></div>
                <div class="error-message" id="recaptchaError" style="display: none;"></div>
            </div>

            <!-- Buttons -->
            <div class="button-group">
                <button type="submit" class="submit-btn" id="submitBtn">{submit_button}</button>
                <button type="reset" class="reset-btn" id="resetBtn">{reset_button}</button>
            </div>
        </form>

        <!-- Success Message -->
        <div class="success-message" id="successMessage">
            {success_message}
        </div>

        <!-- Summary Section -->
        <div class="summary-section" id="summarySection" style="display: none;">
            <h2 class="summary-title">{summary_title}</h2>
            <div id="summaryContent"></div>
        </div>
    </div>

    <script>
        // Define questions for this survey
        window.surveyQuestions = {{
{js_questions}
        }};

        // Define the chapter name for this survey
        window.chapterName = "{chapter_name}";
    </script>

    <!-- EmailJS Library -->
    <script src="https://cdn.jsdelivr.net/npm/emailjs-com@3/dist/email.min.js"></script>
    <script>
        (function() {{
            emailjs.init("wZ_Z4F9Y-8CcFzD2g");
        }})();
    </script>

    <!-- Google reCAPTCHA v2 -->
    <script src="https://www.google.com/recaptcha/api.js" async defer></script>

    <script src="{js_file}"></script>
</body>
</html>
'''

QUESTION_TEMPLATE = '''
            <!-- {q_id_upper} -->
            <div class="question-section">
                <div class="question-title">{question_text}</div>
                <div class="options">
{options_html}
                </div>
            </div>
'''

OPTION_TEMPLATE = '''                    <div class="option">
                        <input type="radio" id="{q_id}_{opt_num}" name="{q_id}" value="{value}"{required}>
                        <label for="{q_id}_{opt_num}">{label}</label>
                    </div>'''


def load_excel_data(excel_path):
    """Load question and UI data from Excel file."""
    wb = load_workbook(excel_path)

    # Load questions
    ws_questions = wb['Questions']
    questions_by_chapter = {}

    for row in ws_questions.iter_rows(min_row=2, values_only=True):
        if not row[0]:  # Skip empty rows
            continue

        chapter_num = int(row[0])
        q_id = row[1]
        q_text = row[2]
        options = [opt for opt in row[3:] if opt]  # Filter out empty options

        if chapter_num not in questions_by_chapter:
            questions_by_chapter[chapter_num] = []

        questions_by_chapter[chapter_num].append({
            'id': q_id,
            'text': q_text,
            'options': options
        })

    # Load UI elements
    ws_ui = wb['UI Elements']
    ui_by_chapter = {}

    for row in ws_ui.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue

        chapter_num = int(row[0])
        ui_by_chapter[chapter_num] = {
            'title': row[1] or '',
            'main_instruction': row[2] or '',
            'conversation_label': row[3] or '',
            'conversation_placeholder': row[4] or '',
            'email_label': row[5] or '',
            'email_placeholder': row[6] or '',
            'submit_button': row[7] or '',
            'reset_button': row[8] or '',
            'success_message': row[9] or '',
            'summary_title': row[10] or ''
        }

    return questions_by_chapter, ui_by_chapter


def generate_question_html(questions):
    """Generate HTML for all questions in a chapter."""
    html_parts = []

    for i, q in enumerate(questions):
        options_html = []
        for opt_idx, opt in enumerate(q['options'], 1):
            required = ' required' if opt_idx == 1 else ''
            opt_html = OPTION_TEMPLATE.format(
                q_id=q['id'],
                opt_num=opt_idx,
                value=opt,
                label=opt,
                required=required
            )
            options_html.append(opt_html)

        q_html = QUESTION_TEMPLATE.format(
            q_id_upper=q['id'].upper(),
            question_text=q['text'],
            options_html='\n'.join(options_html)
        )
        html_parts.append(q_html)

    return ''.join(html_parts)


def generate_js_questions(questions, conversation_label):
    """Generate JavaScript surveyQuestions object content."""
    lines = []
    for q in questions:
        # Escape quotes in question text
        text = q['text'].replace('"', '\\"')
        lines.append(f'            {q["id"]}: "{text}"')

    # Add conversation
    conv_text = conversation_label.replace('"', '\\"')
    lines.append(f'            conversation: "{conv_text}"')

    return ',\n'.join(lines)


def generate_chapter_html(chapter_num, questions, ui_data, lang_config, chapter_name_format):
    """Generate complete HTML for a chapter."""
    questions_html = generate_question_html(questions)
    js_questions = generate_js_questions(questions, ui_data['conversation_label'])

    # Determine conversation section title based on language
    conversation_title = 'Conversen:' if lang_config['html_lang'] == 'es' else 'Discuss:'

    # Format chapter name
    if lang_config['html_lang'] == 'es':
        chapter_name = f"Capítulo {chapter_num}"
    else:
        chapter_name = f"Chapter {chapter_num}"

    html = HTML_TEMPLATE.format(
        html_lang=lang_config['html_lang'],
        title=ui_data['title'],
        css_file=lang_config['css_file'],
        main_instruction=ui_data['main_instruction'],
        questions_html=questions_html,
        conversation_title=conversation_title,
        conversation_label=ui_data['conversation_label'],
        conversation_placeholder=ui_data['conversation_placeholder'],
        email_label=ui_data['email_label'],
        email_placeholder=ui_data['email_placeholder'],
        submit_button=ui_data['submit_button'],
        reset_button=ui_data['reset_button'],
        success_message=ui_data['success_message'],
        summary_title=ui_data['summary_title'],
        js_questions=js_questions,
        chapter_name=chapter_name,
        js_file=lang_config['js_file']
    )

    return html


def main():
    parser = argparse.ArgumentParser(description='Import survey content from Excel to HTML')
    parser.add_argument('--input', '-i', default=DEFAULT_INPUT_FILE,
                        help='Input Excel file path')
    parser.add_argument('--output-dir', '-o', default=HTML_DIR,
                        help='Output directory for HTML files')
    parser.add_argument('--lang', '-l', default='es', choices=['es', 'en'],
                        help='Language code (es=Spanish, en=English)')
    parser.add_argument('--prefix', '-p',
                        help='File prefix (default: capitulo for es, chapter for en)')
    parser.add_argument('--dry-run', '-n', action='store_true',
                        help='Show what would be done without creating files')

    args = parser.parse_args()

    # Get language config
    lang_config = LANG_CONFIG[args.lang].copy()
    if args.prefix:
        lang_config['prefix'] = args.prefix

    print(f"Loading Excel file: {args.input}")
    questions_by_chapter, ui_by_chapter = load_excel_data(args.input)

    print(f"Found {len(questions_by_chapter)} chapters")
    print(f"Language: {args.lang} ({lang_config['html_lang']})")
    print(f"File prefix: {lang_config['prefix']}")
    print(f"Output directory: {args.output_dir}")
    print()

    # Generate HTML files
    for chapter_num in sorted(questions_by_chapter.keys()):
        if chapter_num not in ui_by_chapter:
            print(f"Warning: Chapter {chapter_num} missing UI data, skipping")
            continue

        questions = questions_by_chapter[chapter_num]
        ui_data = ui_by_chapter[chapter_num]

        filename = f"{lang_config['prefix']}_{chapter_num:02d}.html"
        filepath = os.path.join(args.output_dir, filename)

        if args.dry_run:
            print(f"Would create: {filename} ({len(questions)} questions)")
        else:
            html = generate_chapter_html(
                chapter_num, questions, ui_data, lang_config,
                chapter_name_format=lang_config['prefix'].title()
            )

            with open(filepath, 'w', encoding='utf-8') as f:
                f.write(html)

            print(f"Created: {filename} ({len(questions)} questions)")

    if args.dry_run:
        print("\n(Dry run - no files were created)")
    else:
        print(f"\nGenerated {len(questions_by_chapter)} chapter files")


if __name__ == '__main__':
    main()
