#!/usr/bin/env python3
"""
Fix all remaining Spanish text in English chapter files.
Uses Excel translations to replace Spanish options with English.
"""

import os
import re
from openpyxl import load_workbook


def extract_options_mapping(excel_path):
    """Extract Spanish to English option mappings from Excel."""
    wb = load_workbook(excel_path)
    mappings = {}

    for sheet_name in wb.sheetnames:
        cap_match = re.match(r'^cap (\d+)$', sheet_name)
        if not cap_match:
            continue

        chapter_num = int(cap_match.group(1))
        sheet = wb[sheet_name]

        # Extract English options
        english_options = []
        for row in sheet.iter_rows(values_only=True):
            cell = row[0] if row else None
            if cell is not None:
                text = str(cell).strip()
                # Options start with bullets or letters
                if text.startswith('•') or text.startswith('·'):
                    english_options.append(text[1:].strip())
                elif re.match(r'^[A-D]\.\s', text):
                    english_options.append(re.sub(r'^[A-D]\.\s*', '', text).strip())
                elif text in ['Yes', 'No', 'True', 'False', 'Very much', 'Not so much', 'A little bit']:
                    english_options.append(text)

        mappings[chapter_num] = english_options

    return mappings


# Comprehensive Spanish to English translations for chapter 22
CHAPTER_22_TRANSLATIONS = {
    'Mucho': 'Very much',
    'No tanto': 'Not so much',
    'Un poco': 'A little bit',
    'La intimidad no es una actividad simple, sino que involucra nuestras emociones': 'Because intimacy is not a dry activity it involves our emotions',
    'La armonía crea un ciclo de conexión': 'Harmony creates a cycle of connection',
    'Una mejor relación lleva a una intimidad más satisfactoria': 'Better the relationship greater the intimacy',
    'Se debe buscar ayuda profesional': 'One should seek professional help',
    'A veces la abstinencia y la distancia pueden ayudar a generar mayor interés': 'Sometimes abstinence and distance can help generate greater interest',
    'Se necesita comunicar las necesidades y llegar a un entendimiento con la pareja': 'One needs to communicate their needs and reach an understanding with their spouse',
    'Conversen acerca de cómo se sienten cada uno en relación con la intimidad en el matrimonio y qué pueden hacer si es necesario para mejorarla': 'How each one feels regarding the intimacy in the marriage and what can be done if necessary to improve it',
}

# Common translations across chapters
COMMON_TRANSLATIONS = {
    # Yes/No
    'Si': 'Yes',
    'Sí': 'Yes',
    'No': 'No',
    # True/False
    'Verdadero': 'True',
    'Falso': 'False',
    # Common words
    'Dinero': 'Money',
    'Felicidad': 'Happiness',
    'Paz': 'Peace',
    'Placer': 'Pleasure',
    'Amor': 'Love',
    'Autoestima': 'Self-esteem',
    'Riqueza': 'Wealth',
    # Life stages
    'La infancia': 'Childhood',
    'Infancia': 'Childhood',
    'La adolescencia': 'Adolescence',
    'Adolescencia': 'Adolescence',
    'La adultez': 'Adulthood',
    'Adultez': 'Adulthood',
    'Todas las anteriores': 'All of the above',
    'Todo lo anterior': 'All of the above',
    # Chapter-specific common phrases
    'Ninguna': 'None',
    'La relación se vuelve más fácil': 'The relationship becomes easier',
    # Chapter 3
    'Que implican sentirse amados y queridos': 'They entail feeling loved and wanted',
    'Que implican obtención de poder, prestigio y placer': 'They involve gaining power, prestige, and pleasure',
    'La plata te ayuda a conseguir las otras P': 'Money contributes to all of the P\'s',
    'La gente no la busca': 'People don\'t seek it',
    'Lo demás es más importante': 'Other things are more important',
    # Chapter 4
    'Sus logros, apariencia y afecto recibido': 'Her accomplishments, appearance, and received affection',
    'El poder, prestigio y placer que obtiene': 'The power, prestige, and pleasure she obtains',
    'Sentirse querida y amada': 'Feeling loved',
    'Que la provean': 'Being provided for',
    'Sentirse emocionalmente conectada con su esposo': 'Feeling emotionally connected to her husband',
    # Chapter 5
    'A las mujeres les gusta comprar': 'Women like shopping',
    'Tener dinero aumenta su autoestima': 'Having money boosts his self-esteem',
    'Todo el mundo es atraído por el dinero': 'Everyone is enticed by money',
    'No fue suficiente': 'It wasn\'t enough',
    'Quería pasar tiempo con él': 'She wanted to spend time with him',
    'Tienen intereses diferentes': 'They have different interests',
    'No entienden sus diferencias': 'They don\'t understand their differences',
    'Él no la consultó': 'He didn\'t consult with her',
    'Ella quería control': 'She wanted control',
    # Chapter 6
    'Verdadero': 'True',
    'Falso': 'False',
    # Chapter 7
    'Ver su programa de televisión favorito': 'Watch their favorite TV show',
    'Hablar con su pareja': 'Talk with their partner',
    'Reflexionar sobre lo que pasó durante el día': 'Reflect on what happened during the day',
    'Relajarse solo': 'Relax by himself',
    'Así es como se sienten amadas': 'That is how they feel loved',
    'Están aburridas': 'They are bored',
    'Necesitan desahogarse': 'They need to vent',
    'No encuentran interés en los temas de conversación': 'They don\'t find interest in the topics of the conversation',
    'No se dan cuenta de cuánto ella valora hablar': 'They fail to see how much she values talking',
    'Sienten que ya han hablado suficiente': 'They feel they\'ve already spoken enough',
    'Está más feliz': 'She\'s happier',
    'No le afecta': 'It doesn\'t affect her',
    'Se siente sola': 'She feels lonely',
    # UI elements
    'Conversen:': 'Discuss:',
    'Reflexión:': 'Reflection:',
    'Compara tus respuestas': 'Compare your answers',
    'Compartan sus respuestas': 'Share your responses with each other',
}


def fix_chapter_file(filepath, chapter_num):
    """Fix all Spanish text in a chapter file."""
    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()

    original_content = content

    # Apply common translations
    for spanish, english in COMMON_TRANSLATIONS.items():
        # Replace in labels
        content = re.sub(rf'>{re.escape(spanish)}<', f'>{english}<', content)
        # Replace in values
        content = re.sub(rf'value="{re.escape(spanish)}"', f'value="{english}"', content)
        # Replace in text content
        content = content.replace(f'>{spanish}<', f'>{english}<')

    # Apply chapter 22 specific translations
    if chapter_num == 22:
        for spanish, english in CHAPTER_22_TRANSLATIONS.items():
            content = re.sub(rf'>{re.escape(spanish)}<', f'>{english}<', content)
            content = re.sub(rf'value="{re.escape(spanish)}"', f'value="{english}"', content)

    if content != original_content:
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(content)
        return True
    return False


def main():
    for chapter_num in range(1, 23):
        filepath = f'chapter_{chapter_num:02d}.html'
        if os.path.exists(filepath):
            if fix_chapter_file(filepath, chapter_num):
                print(f"Fixed {filepath}")
            else:
                print(f"No changes needed for {filepath}")


if __name__ == '__main__':
    main()
